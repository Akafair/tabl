# -*- coding: cp950 -*-

from openpyxl import *
from openpyxl.compat import range
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from decimal import Decimal

__author__ = 'Tom Lin'

# Excel Format
header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(fill_type='solid', start_color='0070C0', end_color='0070C0')
alignment = Alignment(horizontal='center', vertical='center')
default_font = Font(name='Calibri', size=10, bold=False)
border_thin = Border(left=Side(border_style='thin', color='000000'),
                     right=Side(border_style='thin', color='000000'),
                     top=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000'))
border_medium = Border(left=Side(border_style='medium', color='000000'),
                       right=Side(border_style='medium', color='000000'),
                       top=Side(border_style='medium', color='000000'),
                       bottom=Side(border_style='medium', color='000000'))
fail_fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')  # Red
pass_fill = PatternFill(fill_type='solid', start_color='00B050', end_color='00B050')  # Light Green
fail_v_fill = PatternFill(fill_type='solid', start_color='0070C0', end_color='0070C0')
fail_l_fill = PatternFill(fill_type='solid', start_color='FFC000', end_color='FFC000')


def write_excel_row(ws, items, row, col):

    for i in items.split(','):
        if re.search(r'[a-zA-Z%]', i):
            ws.cell(row=row, column=col).value = i
        elif re.search(r'^\d+|^-\d+', i):
            if re.search(r'\.', i):
                ws.cell(row=row, column=col).value = Decimal(i)
                # ws.cell(row=row, column=col).number_format = '0.00'
            else:
                ws.cell(row=row, column=col).value = int(i)
        else:
            ws.cell(row=row, column=col).value = i

        ws.cell(row=row, column=col).font = default_font
        ws.cell(row=row, column=col).alignment = alignment
        col += 1


def write_excel_header(ws, header, row, col):

    for i in header.split(','):
        ws.cell(row=row, column=col).value = i
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).alignment = alignment
        ws.cell(row=row, column=col).fill = header_fill
        col += 1
